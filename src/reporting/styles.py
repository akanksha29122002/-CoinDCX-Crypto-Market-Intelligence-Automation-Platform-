from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- BRANDING PALETTE (CoinDCX Corporate Visual System) ---
COLOR_PRIMARY_NAVY = "0F172A"   # Deep slate navy blue for headers
COLOR_SECONDARY_ORANGE = "FF5A00" # CoinDCX Brand Orange for accent lines, main highlights
COLOR_ZEBRA_LIGHT = "F8FAFC"      # Pale slate for clean alternating rows
COLOR_WHITE = "FFFFFF"
COLOR_BORDER_GREY = "E2E8F0"     # Soft modern borders

COLOR_ALERT_GREEN_FILL = "D1FAE5"  # Soft positive background
COLOR_ALERT_GREEN_TXT = "065F46"
COLOR_ALERT_RED_FILL = "FFE4E6"    # Soft crash background
COLOR_ALERT_RED_TXT = "991B1B"
COLOR_ALERT_GOLD_FILL = "FEF3C7"   # Soft warning background
COLOR_ALERT_GOLD_TXT = "92400E"

# --- SYSTEM FONTS (Google Fonts Inter Styling Equivalents) ---
FONT_TITLE = Font(name="Segoe UI", size=16, bold=True, color=COLOR_PRIMARY_NAVY)
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="64748B")
FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color=COLOR_WHITE)
FONT_ACCENT_HEADER = Font(name="Segoe UI", size=11, bold=True, color=COLOR_SECONDARY_ORANGE)
FONT_DATA_BOLD = Font(name="Segoe UI", size=10, bold=True, color=COLOR_PRIMARY_NAVY)
FONT_DATA_REGULAR = Font(name="Segoe UI", size=10, color="1E293B")
FONT_MUTED = Font(name="Segoe UI", size=9, color="64748B")

# --- ALIGNMENT PRESETS ---
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# --- CELL FILL STYLES ---
FILL_HEADER = PatternFill(start_color=COLOR_PRIMARY_NAVY, end_color=COLOR_PRIMARY_NAVY, fill_type="solid")
FILL_ACCENT = PatternFill(start_color="FFF3EA", end_color="FFF3EA", fill_type="solid") # Very soft orange fill
FILL_ZEBRA = PatternFill(start_color=COLOR_ZEBRA_LIGHT, end_color=COLOR_ZEBRA_LIGHT, fill_type="solid")
FILL_WHITE = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

FILL_ALERT_GREEN = PatternFill(start_color=COLOR_ALERT_GREEN_FILL, end_color=COLOR_ALERT_GREEN_FILL, fill_type="solid")
FILL_ALERT_RED = PatternFill(start_color=COLOR_ALERT_RED_FILL, end_color=COLOR_ALERT_RED_FILL, fill_type="solid")
FILL_ALERT_GOLD = PatternFill(start_color=COLOR_ALERT_GOLD_FILL, end_color=COLOR_ALERT_GOLD_FILL, fill_type="solid")

# --- BORDERS ---
border_side_thin = Side(border_style="thin", color=COLOR_BORDER_GREY)
BORDER_ALL_THIN = Border(left=border_side_thin, right=border_side_thin, top=border_side_thin, bottom=border_side_thin)
BORDER_TOP_BOTTOM = Border(top=border_side_thin, bottom=Side(border_style="medium", color=COLOR_PRIMARY_NAVY))
BORDER_TOTAL_ROW = Border(top=Side(border_style="thin", color=COLOR_PRIMARY_NAVY), bottom=Side(border_style="double", color=COLOR_PRIMARY_NAVY))
