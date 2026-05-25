import os
import logging
from datetime import datetime
import pandas as pd
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from src.database.queries import get_top_movers, get_historical_volatility_stats
from src.database.models import MarketOHLCVHourly, Asset
from src.reporting.styles import (
    FONT_TITLE, FONT_SUBTITLE, FONT_HEADER, FONT_ACCENT_HEADER,
    FONT_DATA_REGULAR, FONT_DATA_BOLD, FONT_MUTED,
    ALIGN_LEFT, ALIGN_RIGHT, ALIGN_CENTER, ALIGN_HEADER,
    FILL_HEADER, FILL_ACCENT, FILL_ZEBRA, FILL_WHITE,
    FILL_ALERT_RED, FILL_ALERT_GREEN, FILL_ALERT_GOLD,
    BORDER_ALL_THIN, BORDER_TOTAL_ROW, BORDER_TOP_BOTTOM,
    COLOR_PRIMARY_NAVY, COLOR_SECONDARY_ORANGE
)

logger = logging.getLogger(__name__)

def generate_daily_report(session: Session, output_dir: str) -> str:
    """
    Orchestrates the programmatical extraction of timeseries records, 
    compiling an executive-grade Excel Report using openpyxl.
    """
    logger.info("Initializing Daily Crypto Report Generator Engine...")
    os.makedirs(output_dir, exist_ok=True)
    report_filename = f"Daily_Crypto_Report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(output_dir, report_filename)

    wb = Workbook()
    
    # 1. GENERATE RAW DATA TAB (Data Source)
    ws_raw = wb.active
    ws_raw.title = "Raw_Market_Ticks"
    _populate_raw_data_sheet(session, ws_raw)

    # 2. GENERATE MARKET ANALYTICS SUMMARY TAB (Pivot / Indicator Engine)
    ws_analytics = wb.create_sheet(title="Market_Analytics")
    _populate_analytics_sheet(session, ws_analytics)

    # 3. GENERATE THE EXECUTIVE CONTROL DASHBOARD TAB
    ws_dashboard = wb.create_sheet(title="Executive_Dashboard", index=0)
    _populate_dashboard_sheet(session, ws_dashboard)

    # Auto-adjust column widths across all sheets to prevent visual truncation
    for ws in wb.worksheets:
        _auto_fit_columns(ws)

    wb.save(output_path)
    logger.info(f"Successfully generated styled Excel report at: {output_path}")
    
    # Create a symlink or secondary copy as the default static name 'Daily_Crypto_Report.xlsx'
    default_path = os.path.join(output_dir, "Daily_Crypto_Report.xlsx")
    try:
        if os.path.exists(default_path):
            os.remove(default_path)
        wb.save(default_path)
        logger.info(f"Updated default static spreadsheet copy: {default_path}")
    except Exception as e:
        logger.warning(f"Could not update static default copy link: {e}")

    return output_path

def _populate_raw_data_sheet(session: Session, ws) -> None:
    """
    Extracts chronological OHLCV timeseries directly from PostgreSQL, 
    populating a flat database worksheet.
    """
    ws.views.sheetView[0].showGridLines = True
    
    headers = ["Symbol", "Timestamp", "Open ($)", "High ($)", "Low ($)", "Close ($)", "Volume", "Hourly Volatility"]
    
    # Style Header Row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_ALL_THIN

    # Query all ticks for past 7 days
    ticks = session.query(MarketOHLCVHourly).order_index = MarketOHLCVHourly.timestamp.desc()
    ticks = ticks.limit(500).all() # Cap for readability

    # Write Data rows with zebra coloring
    for r_idx, tick in enumerate(ticks, 2):
        row_data = [
            tick.symbol,
            tick.timestamp.strftime("%Y-%m-%d %H:%M"),
            float(tick.open),
            float(tick.high),
            float(tick.low),
            float(tick.close),
            float(tick.volume),
            float(tick.hourly_volatility) if tick.hourly_volatility else 0.0
        ]
        
        fill = FILL_ZEBRA if r_idx % 2 == 0 else FILL_WHITE
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_DATA_REGULAR
            cell.fill = fill
            cell.border = BORDER_ALL_THIN
            
            # Format columns
            if c_idx in [3, 4, 5, 6]:
                cell.number_format = "$#,##0.00"
                cell.alignment = ALIGN_RIGHT
            elif c_idx == 7:
                cell.number_format = "#,##0.00"
                cell.alignment = ALIGN_RIGHT
            elif c_idx == 8:
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_CENTER

def _populate_analytics_sheet(session: Session, ws) -> None:
    """
    Compiles structured Pivot-like analytical aggregates 
    and applies programmatic conditional formatting.
    """
    ws.views.sheetView[0].showGridLines = True
    
    # Main title
    ws.cell(row=1, column=1, value="CoinDCX Market Intelligence Engine").font = FONT_TITLE
    ws.cell(row=2, column=1, value="Market Indicators & Anomaly Performance Summary").font = FONT_SUBTITLE
    
    headers = ["Symbol", "Name", "Market Cap ($)", "24h Close ($)", "24h Price Change (%)", "24h Volatility (%)", "24h Rolling SMA ($)"]
    
    # Style Header Row
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_ALL_THIN

    # Query all assets
    assets = session.query(Asset).all()
    
    # Gather data & populate rows
    r_idx = header_row + 1
    for asset in assets:
        # Calculate moving variables via SQL logic
        stats = get_historical_volatility_stats(session, asset.symbol, limit=24)
        
        if not stats:
            continue
            
        latest_tick = stats[0]
        earliest_tick = stats[-1]
        
        close_24h = latest_tick["close"]
        price_change_pct = ((close_24h - earliest_tick["close"]) / earliest_tick["close"]) * 100.0 if earliest_tick["close"] > 0 else 0.0
        
        # Volatility is average of hourly returns std dev
        volatility = sum(s["rolling_std_7d"] for s in stats) / len(stats) if stats else 0.0
        sma_24h = latest_tick["sma_24h"]

        row_data = [
            asset.symbol,
            asset.name,
            float(asset.current_market_cap) if asset.current_market_cap else 0.0,
            close_24h,
            price_change_pct / 100.0,  # Programmed as raw ratio for percentage formatting
            volatility,
            sma_24h
        ]
        
        fill = FILL_ZEBRA if r_idx % 2 == 0 else FILL_WHITE
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_DATA_REGULAR
            cell.fill = fill
            cell.border = BORDER_ALL_THIN
            
            # Formats and Conditional Alerts
            if c_idx == 3:
                cell.number_format = "$#,##0.00"
                cell.alignment = ALIGN_RIGHT
            elif c_idx == 4 or c_idx == 7:
                cell.number_format = "$#,##0.00"
                cell.alignment = ALIGN_RIGHT
            elif c_idx == 5:
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_RIGHT
                # Inline conditional format color rules
                if val <= -0.05: # Price Crash Alert
                    cell.fill = FILL_ALERT_RED
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
                elif val >= 0.05: # Positive Trend Alert
                    cell.fill = FILL_ALERT_GREEN
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
            elif c_idx == 6:
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_RIGHT
                if val > 0.08: # Excessive Volatility Alert
                    cell.fill = FILL_ALERT_GOLD
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="92400E")
            elif c_idx == 1:
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_DATA_BOLD
            else:
                cell.alignment = ALIGN_LEFT
                
        r_idx += 1

def _populate_dashboard_sheet(session: Session, ws) -> None:
    """
    Constructs the master executive control panel, embedding 
    KPI summaries, validation drop-downs, and dynamic XLOOKUP rows.
    """
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Main Title Header Panel
    ws.cell(row=1, column=1, value="CoinDCX Crypto Market Intelligence & Automation Platform").font = FONT_TITLE
    ws.cell(row=2, column=1, value="Production-Grade Operations & Risk Management Dashboard").font = FONT_SUBTITLE
    
    # Accent line
    for col in range(1, 9):
        ws.cell(row=3, column=col).border = Border(bottom=Side(style="medium", color=COLOR_SECONDARY_ORANGE))

    # 2. KPI Cards Block
    _draw_kpi_card(ws, start_row=5, start_col=1, title="ACTIVE TRACKED ASSETS", value=str(session.query(Asset).count()))
    
    # Calculate aggregate sums
    total_cap = session.query(Asset.current_market_cap).all()
    sum_cap = sum(float(c[0]) for c in total_cap if c[0])
    _draw_kpi_card(ws, start_row=5, start_col=3, title="AGGREGATE MARKET CAP", value=f"${sum_cap:,.2f}")
    
    # 24h Vol Movers
    movers = get_top_movers(session, hours=24)
    movers_val = f"{movers[0][0]} (+{movers[0][3]}%)" if movers else "N/A"
    _draw_kpi_card(ws, start_row=5, start_col=6, title="24H VOLATILITY LEADER", value=movers_val)

    # 3. INTERACTIVE LOOKUP INTERFACE (With formulas!)
    ws.cell(row=10, column=1, value="🔍 Dynamic Operations Search Terminal").font = Font(name="Segoe UI", size=12, bold=True, color=COLOR_PRIMARY_NAVY)
    ws.cell(row=11, column=1, value="Select a target symbol from the drop-down menu to evaluate active status.").font = FONT_SUBTITLE

    # Input Cell
    ws.cell(row=13, column=1, value="Select Asset Symbol:").font = FONT_DATA_BOLD
    ws.cell(row=13, column=1).alignment = ALIGN_LEFT
    
    # Target Dropdown
    search_cell = ws.cell(row=13, column=2, value="BTC")
    search_cell.font = Font(name="Segoe UI", size=11, bold=True, color=COLOR_SECONDARY_ORANGE)
    search_cell.alignment = ALIGN_CENTER
    search_cell.border = BORDER_ALL_THIN
    search_cell.fill = FILL_ACCENT

    # Setup Pydantic-like Data Validation list rules inside Excel
    assets_symbols = [a.symbol for a in session.query(Asset.symbol).all()]
    if assets_symbols:
        formula_str = f'"{",".join(assets_symbols)}"'
        dv = DataValidation(type="list", formula1=formula_str, allow_blank=True)
        dv.error = 'Your input value is invalid. Please select an active tracked asset.'
        dv.errorTitle = 'Invalid Asset Code'
        dv.prompt = 'Choose an asset code from the list.'
        dv.promptTitle = 'Select Token Code'
        ws.add_data_validation(dv)
        dv.add(search_cell)

    # Dynamic Lookup Rows using standard lookup mapping formulas (XLOOKUP)
    lookup_headers = ["Asset Name", "Current Close Price ($)", "Historical Volatility Index", "24h Price Return"]
    for idx, header in enumerate(lookup_headers, 1):
        ws.cell(row=15, column=idx, value=header).font = FONT_HEADER
        ws.cell(row=15, column=idx).fill = FILL_HEADER
        ws.cell(row=15, column=idx).alignment = ALIGN_HEADER
        ws.cell(row=15, column=idx).border = BORDER_ALL_THIN

    # Implement standard lookup formulas pointing to the Analytics sheet
    # Row 16 details:
    # Asset Name:   XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!B:B, "NOT FOUND")
    # Close Price:  XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!D:D, 0.0)
    # Volatility:   XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!F:F, 0.0)
    # 24h Return:   XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!E:E, 0.0)

    cell_name = ws.cell(row=16, column=1, value='=XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!B:B, "NOT FOUND")')
    cell_price = ws.cell(row=16, column=2, value='=XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!D:D, 0.0)')
    cell_vol = ws.cell(row=16, column=3, value='=XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!F:F, 0.0)')
    cell_ret = ws.cell(row=16, column=4, value='=XLOOKUP(B13, Market_Analytics!A:A, Market_Analytics!E:E, 0.0)')

    for idx, cell in enumerate([cell_name, cell_price, cell_vol, cell_ret], 1):
        cell.font = FONT_DATA_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL_THIN
        cell.fill = FILL_ZEBRA
        
        if idx == 2:
            cell.number_format = "$#,##0.00"
        elif idx in [3, 4]:
            cell.number_format = "0.00%"

    # 4. EMBED NATIVE LINE CHART
    # Plot historical hourly ticks of the raw data sheet
    chart = LineChart()
    chart.title = "Hourly Market Price Matrix (Chronological Trend Line)"
    chart.style = 13
    chart.y_axis.title = "Asset Price ($)"
    chart.x_axis.title = "Hourly Interval Index"
    chart.width = 18
    chart.height = 10

    # Data is columns 3 (Open) to 6 (Close) from Raw_Market_Ticks sheet
    data = Reference(ws_raw = ws.parent["Raw_Market_Ticks"], min_col=6, min_row=1, max_row=100) # Plot last 100 close ticks
    chart.add_data(data, titles_from_data=True)

    ws.add_chart(chart, "A19")

def _draw_kpi_card(ws, start_row: int, start_col: int, title: str, value: str) -> None:
    """
    Renders styled rectangular KPI indicator cards in Excel grids.
    """
    # Merge cells for card layout
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+1)
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+2, end_column=start_col+1)

    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = FONT_MUTED
    title_cell.alignment = ALIGN_CENTER
    title_cell.fill = FILL_ZEBRA

    val_cell = ws.cell(row=start_row+1, column=start_col, value=value)
    val_cell.font = FONT_ACCENT_HEADER if "Leader" in title or "+" in value else FONT_DATA_BOLD
    val_cell.font = Font(name="Segoe UI", size=13, bold=True, color=COLOR_PRIMARY_NAVY)
    val_cell.alignment = ALIGN_CENTER
    val_cell.fill = FILL_ZEBRA

    # Border framing
    thin_border = Side(style='thin', color='CBD5E1')
    for r in range(start_row, start_row+3):
        for c in range(start_col, start_col+2):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=thin_border if c == start_col else None,
                right=thin_border if c == start_col+1 else None,
                top=thin_border if r == start_row else None,
                bottom=thin_border if r == start_row+2 else None
            )

def _auto_fit_columns(ws) -> None:
    """
    Dynamically resizes sheet columns based on standard string lengths.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            # Ignore long formula codes to avoid oversized column widths
            if val.startswith("="):
                max_len = max(max_len, 15)
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
